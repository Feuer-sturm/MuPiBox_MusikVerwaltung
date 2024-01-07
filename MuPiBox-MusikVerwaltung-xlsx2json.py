import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUTFILE = 'MuPiBox_MusikVerwaltung_Template.xlsx'
OUTPUTFILE = 'data.json'

wb = load_workbook(filename=INPUTFILE)
ws = wb.active
my_list = []

FIRST_COLUMN = 2    #entspricht Spalte B
LAST_COLUM = 16     #entspricht Spalte P
FIRST_ROW = 2       #Zeile in Excel, wo die Attribute beginnen
last_row = len(list(ws.rows))
index = 0

for row in range(FIRST_ROW, last_row + 1):                      #Starte ab Zeile 2, Zeile 1 enthält Erklärungen zu den Feldern
    my_dict = {}
    if ws['A' + str(row)].value == 'x':                         #Es wird geprüft ob der Eintrag welcher ein "x" enthält zur json Datei hinzugefügt werden soll 
        my_dict["index"] = index
        for column in range(FIRST_COLUMN, LAST_COLUM + 1):
            column_letter = get_column_letter(column)
            if row > FIRST_ROW and ws[column_letter + str(row)].value is not None :
                my_dict[ws[column_letter + str(FIRST_ROW)].value] = ws[column_letter + str(row)].value
        my_list.append(my_dict)
        index = index + 1

data = json.dumps(my_list, sort_keys=False, indent=4, ensure_ascii=False)
with open(OUTPUTFILE, 'w', encoding='utf-8') as f:
    f.write(data)
